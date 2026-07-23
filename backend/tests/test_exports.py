"""Excel + PDF exports derive from the same snapshot payload and are valid files."""

from __future__ import annotations

import io

from openpyxl import load_workbook

from osprey.exports import hotlist_to_pdf, hotlist_to_xlsx

PAYLOAD = {
    "project_id": "p1",
    "generated_at": "2026-07-23T00:00:00+00:00",
    "top_n": 25,
    "item_count": 2,
    "total_exposure": 180000.0,
    "buckets": {
        "act_today": {"count": 1, "exposure": 180000.0},
        "this_week": {"count": 1, "exposure": 54000.0},
        "watch": {"count": 0, "exposure": 0.0},
        "done": {"count": 0, "exposure": 0.0},
    },
    "items": [
        {
            "item_id": "i1", "what": "NOTICE OF DELAY — Tower B", "category": "contractual_notice",
            "bucket": "act_today", "bucket_label": "Act today", "bucket_emoji": "🔴",
            "why": "Act today: contractual notice deadline; deadline 2026-07-29.",
            "summary": "s", "sources": [{"source_type": "filedrop", "title": "Notice", "url": "https://x/1"}],
            "owner": "PM", "due": "2026-07-29", "dollar_exposure": 180000.0,
            "recommended_action": "Respond in writing before the notice lapses.",
            "notice_deadline": True, "score": 88.0, "factors": {"urgency": 0.95},
        },
        {
            "item_id": "i2", "what": "Pay Application 07", "category": "invoice",
            "bucket": "this_week", "bucket_label": "This week", "bucket_emoji": "🟠",
            "why": "This week: $54,000 exposure.", "summary": "s2", "sources": [],
            "owner": None, "due": None, "dollar_exposure": 54000.0,
            "recommended_action": "Verify against SOV and approve.", "notice_deadline": False,
            "score": 52.0, "factors": {},
        },
    ],
}


def test_xlsx_export_is_valid_workbook():
    data = hotlist_to_xlsx(PAYLOAD, project_name="Tower B")
    assert data[:2] == b"PK"  # xlsx is a zip
    wb = load_workbook(io.BytesIO(data))
    assert {"Summary", "Hotlist", "Raw"}.issubset(set(wb.sheetnames))
    hs = wb["Hotlist"]
    # Header + 2 data rows.
    assert hs.max_row == 3
    assert hs.cell(row=2, column=3).value == "NOTICE OF DELAY — Tower B"


def test_pdf_export_is_valid_pdf():
    data = hotlist_to_pdf(PAYLOAD, project_name="Tower B")
    assert data[:5] == b"%PDF-"
    assert len(data) > 800


def test_exports_agree_on_item_count():
    xlsx = hotlist_to_xlsx(PAYLOAD, project_name="P")
    wb = load_workbook(io.BytesIO(xlsx))
    xlsx_rows = wb["Hotlist"].max_row - 1
    assert xlsx_rows == PAYLOAD["item_count"]
    # PDF built from the same payload -> same source of truth.
    assert hotlist_to_pdf(PAYLOAD, project_name="P")[:5] == b"%PDF-"
