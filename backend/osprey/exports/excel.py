"""Excel export (openpyxl) — Summary + Hotlist + Raw sheets, brand-styled."""

from __future__ import annotations

import io
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Brand tokens (BRAND.md)
INK = "0E1A2B"
EMBER = "FF6A2B"
MIST = "F6F7F9"
WHITE = "FFFFFF"
BUCKET_FILL = {
    "act_today": "E5484D",
    "this_week": "F5A623",
    "watch": "EAB308",
    "done": "30A46C",
}
_THIN = Side(style="thin", color="E4E7EC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header(cell, text: str) -> None:
    cell.value = text
    cell.font = Font(bold=True, color=WHITE, name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=INK)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = _BORDER


def hotlist_to_xlsx(payload: dict[str, Any], *, project_name: str = "Project") -> bytes:
    wb = Workbook()

    # ---- Summary ----------------------------------------------------------- #
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Osprey — Hotlist"
    ws["A1"].font = Font(bold=True, size=18, color=INK)
    ws["A2"] = "The foreman that never sleeps."
    ws["A2"].font = Font(italic=True, size=10, color="667085")
    ws["A4"] = "Project"
    ws["B4"] = project_name
    ws["A5"] = "Generated"
    ws["B5"] = payload.get("generated_at", "")
    ws["A6"] = "Items"
    ws["B6"] = payload.get("item_count", 0)
    ws["A7"] = "Total $ exposure"
    ws["B7"] = payload.get("total_exposure", 0)
    ws["B7"].number_format = "$#,##0"
    for r in range(4, 8):
        ws[f"A{r}"].font = Font(bold=True, color=INK)

    ws["A9"] = "Bucket"
    ws["B9"] = "Count"
    ws["C9"] = "$ Exposure"
    for col in ("A9", "B9", "C9"):
        _header(ws[col], ws[col].value)
    buckets = payload.get("buckets", {})
    order = ["act_today", "this_week", "watch", "done"]
    labels = {
        "act_today": "🔴 Act today",
        "this_week": "🟠 This week",
        "watch": "🟡 Watch",
        "done": "✅ Done",
    }
    row = 10
    for key in order:
        b = buckets.get(key, {"count": 0, "exposure": 0.0})
        ws.cell(row=row, column=1, value=labels[key])
        ws.cell(row=row, column=2, value=b.get("count", 0))
        c = ws.cell(row=row, column=3, value=round(b.get("exposure", 0.0), 2))
        c.number_format = "$#,##0"
        fill = PatternFill("solid", fgColor=BUCKET_FILL[key])
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).font = Font(bold=True, color=WHITE)
        row += 1
    for col, width in {"A": 22, "B": 12, "C": 16}.items():
        ws.column_dimensions[col].width = width

    # ---- Hotlist ----------------------------------------------------------- #
    hs = wb.create_sheet("Hotlist")
    cols = [
        ("Rank", 6),
        ("Bucket", 14),
        ("What", 40),
        ("Category", 16),
        ("Why", 50),
        ("Owner", 16),
        ("Due", 14),
        ("$ Exposure", 14),
        ("Recommended action", 46),
        ("Score", 8),
        ("Source", 40),
    ]
    for idx, (name, width) in enumerate(cols, start=1):
        _header(hs.cell(row=1, column=idx), name)
        hs.column_dimensions[get_column_letter(idx)].width = width
    hs.freeze_panes = "A2"

    for i, item in enumerate(payload.get("items", []), start=1):
        r = i + 1
        bucket = item.get("bucket", "watch")
        vals = [
            i,
            item.get("bucket_label", bucket),
            item.get("what", ""),
            item.get("category", ""),
            item.get("why", ""),
            item.get("owner") or "",
            item.get("due") or "",
            item.get("dollar_exposure"),
            item.get("recommended_action", ""),
            item.get("score", 0),
            "",
        ]
        for c_idx, v in enumerate(vals, start=1):
            cell = hs.cell(row=r, column=c_idx, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=c_idx in (3, 5, 9))
            cell.border = _BORDER
        hs.cell(row=r, column=8).number_format = "$#,##0"
        # Bucket cell fill + first source hyperlink.
        bcell = hs.cell(row=r, column=2)
        bcell.fill = PatternFill("solid", fgColor=BUCKET_FILL.get(bucket, "EAB308"))
        bcell.font = Font(bold=True, color=WHITE)
        sources = item.get("sources", [])
        src_cell = hs.cell(row=r, column=11)
        if sources:
            first = sources[0]
            label = f"{first.get('source_type', '')}: {first.get('title', '')}"[:80]
            src_cell.value = label
            if first.get("url"):
                src_cell.hyperlink = first["url"]
                src_cell.font = Font(color=EMBER, underline="single")
    if payload.get("items"):
        hs.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(payload['items']) + 1}"

    # ---- Raw (audit) ------------------------------------------------------- #
    raw = wb.create_sheet("Raw")
    _header(raw.cell(row=1, column=1), "item_id")
    _header(raw.cell(row=1, column=2), "factors (json)")
    raw.column_dimensions["A"].width = 36
    raw.column_dimensions["B"].width = 120
    for i, item in enumerate(payload.get("items", []), start=2):
        raw.cell(row=i, column=1, value=item.get("item_id", ""))
        raw.cell(row=i, column=2, value=json.dumps(item.get("factors", {}), default=str))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
